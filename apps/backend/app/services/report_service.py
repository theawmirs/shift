import sqlite3
import datetime
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.core.config import settings
from app.core import jalali
from app.services import record_service, leave_service
from app.db.database import DBAdapter

def current_month_key() -> str:
    s = record_service.today_str()
    return s[:7]

def month_days(jy: int, jm: int) -> list[str]:
    days_in_month = 31 if jm <= 6 else (30 if jm <= 11 else 29)
    if jm == 12:
        try:
            gy, gm, gd = jalali.jalali_to_gregorian(jy, 12, 30)
            jy2, jm2, jd2 = jalali.gregorian_to_jalali(gy, gm, gd)
            if (jy2, jm2, jd2) == (jy, 12, 30):
                days_in_month = 30
        except Exception:
            pass
    return [jalali.jalali_date_str(jy, jm, d) for d in range(1, days_in_month + 1)]

def compute_month(conn: DBAdapter, month_key: str, user_id: int | None = None) -> dict:
    parts = month_key.split("-")
    jy, jm = int(parts[0]), int(parts[1])
    m_name = jalali.MONTHS_FA[jm - 1]
    days = month_days(jy, jm)

    # Batch fetch all data for the month in 4 queries instead of 300+ round-trips
    from app.db.schema import get_user_settings
    u_settings = get_user_settings(conn, user_id=user_id)
    standard = float(u_settings.get("standard_hours", "8"))
    start_time_end_str = u_settings.get("start_time_end", "09:15")
    end_time_end_str = u_settings.get("end_time_end", "17:15")
    default_work_mode = u_settings.get("default_work_mode", "office")

    # Fetch holidays for month
    holiday_rows = conn.execute("SELECT date, name FROM holidays WHERE substr(date,1,7)=?", (month_key,)).fetchall()
    holiday_map = {r["date"]: r["name"] for r in holiday_rows}

    # Fetch work modes for month
    if user_id is None:
        wm_rows = conn.execute("SELECT shamsi_date, mode FROM day_work_mode WHERE substr(shamsi_date,1,7)=? AND user_id IS NULL", (month_key,)).fetchall()
    else:
        wm_rows = conn.execute("SELECT shamsi_date, mode FROM day_work_mode WHERE substr(shamsi_date,1,7)=? AND user_id=?", (month_key, user_id)).fetchall()
    work_mode_map = {r["shamsi_date"]: r["mode"] for r in wm_rows}

    # Fetch all events for month
    if user_id is None:
        event_rows = conn.execute("SELECT shamsi_date, event_type, ts_utc, note FROM events WHERE substr(shamsi_date,1,7)=? AND user_id IS NULL ORDER BY ts_utc ASC", (month_key,)).fetchall()
    else:
        event_rows = conn.execute("SELECT shamsi_date, event_type, ts_utc, note FROM events WHERE substr(shamsi_date,1,7)=? AND user_id=? ORDER BY ts_utc ASC", (month_key, user_id)).fetchall()

    events_by_date = {}
    for r in event_rows:
        dt = datetime.datetime.fromisoformat(r["ts_utc"])
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        # CRITICAL: Always convert UTC timestamp to Iran / Tehran local time for accurate display
        dt_tehran = dt.astimezone(settings.tehran_tz)
        events_by_date.setdefault(r["shamsi_date"], []).append((r["event_type"], dt_tehran, r["note"]))

    # Fetch daily leaves for user in this month
    if user_id is None:
        dl_rows = conn.execute("SELECT id, user_id, start_date, end_date, type, reason, hours, created_at FROM daily_leaves WHERE user_id IS NULL AND substr(start_date,1,7) <= ? AND substr(end_date,1,7) >= ?", (month_key, month_key)).fetchall()
    else:
        dl_rows = conn.execute("SELECT id, user_id, start_date, end_date, type, reason, hours, created_at FROM daily_leaves WHERE user_id=? AND substr(start_date,1,7) <= ? AND substr(end_date,1,7) >= ?", (user_id, month_key, month_key)).fetchall()

    sh_end, sm_end = record_service.parse_time_str(start_time_end_str, "09:15")
    eh_end, em_end = record_service.parse_time_str(end_time_end_str, "17:15")

    rows = []
    total_net = 0.0
    total_gross = 0.0
    total_leave = 0.0
    total_ot = 0.0
    total_deficit = 0.0
    total_late = 0.0
    late_days = 0
    work_days = 0
    holiday_days = 0
    holiday_worked = 0
    remote_days = 0

    for sdate in days:
        jy_d, jm_d, jd_d = record_service.parse_date(sdate)
        gy, gm, gd = jalali.jalali_to_gregorian(jy_d, jm_d, jd_d)
        wdf = jalali.weekday_fa(gy, gm, gd)
        is_friday = (datetime.date(gy, gm, gd).weekday() == 4)
        hol_name = holiday_map.get(sdate)
        is_hol = is_friday or bool(hol_name)
        wm = work_mode_map.get(sdate, default_work_mode)
        events = events_by_date.get(sdate, [])

        in_dt = None
        out_dt = None
        leave_intervals = []
        leave_open = False
        cur_ls = None

        for et, dt, _ in events:
            if et == "in" and in_dt is None:
                in_dt = dt
            elif et == "out":
                out_dt = dt
            elif et == "leave_start":
                cur_ls = dt
                leave_open = True
            elif et == "leave_end" and cur_ls is not None:
                leave_intervals.append((cur_ls, dt))
                cur_ls = None
                leave_open = False

        gross = 0.0
        if in_dt and out_dt:
            gross = max(0.0, (record_service.company_clock(out_dt) - record_service.company_clock(in_dt)).total_seconds() / 3600.0)

        leave_closed = sum((record_service.company_clock(b) - record_service.company_clock(a)).total_seconds() / 3600.0 for a, b in leave_intervals)
        leave_h = max(0.0, leave_closed)

        overtime = 0.0
        ot_declared = False
        for et, _, note in events:
            if et == "out" and note and "ot:" in note:
                try:
                    overtime = float(note.split("ot:")[1].strip())
                    ot_declared = True
                except Exception:
                    pass

        late = 0.0
        if in_dt and not is_hol:
            in_clock = record_service.company_clock(in_dt)
            lim_in = in_clock.replace(hour=sh_end, minute=sm_end, second=0, microsecond=0)
            if in_clock > lim_in:
                late = (in_clock - lim_in).total_seconds() / 3600.0

        if is_hol:
            net = max(0.0, gross - leave_h)
            overtime = net
            deficit = 0.0
            ot_declared = True
        else:
            if ot_declared and overtime > 0:
                net = max(0.0, gross - leave_h)
            elif in_dt and out_dt:
                effective_presence = max(0.0, gross - leave_h)
                net = min(standard, effective_presence)
            else:
                net = max(0.0, gross - leave_h)

            deficit = max(0.0, standard - net) if (in_dt or out_dt) else 0.0

        has_events = bool(events)
        d_item = {
            "date": sdate,
            "year": jy_d,
            "month": jm_d,
            "day": jd_d,
            "weekday": wdf,
            "is_holiday": is_hol,
            "holiday_name": hol_name,
            "has_events": has_events,
            "in": in_dt.strftime("%H:%M") if in_dt else None,
            "out": out_dt.strftime("%H:%M") if out_dt else None,
            "leave_intervals": [[a.strftime("%H:%M"), b.strftime("%H:%M")] for a, b in leave_intervals],
            "leave_open": leave_open,
            "gross": round(gross, 2),
            "leave": round(leave_h, 2),
            "net": round(net, 2),
            "late": round(late, 2),
            "deficit": round(deficit, 2),
            "overtime": round(overtime, 2),
            "ot_declared": ot_declared,
            "work_mode": wm,
            "work_mode_label": record_service.WORK_MODE_LABEL.get(wm, wm),
            "day_status": "idle",
            "day_status_label": "آماده",
            "day_status_reason": None,
        }

        if is_hol:
            holiday_days += 1
            if has_events:
                holiday_worked += 1
        else:
            if has_events:
                work_days += 1

        if has_events:
            rows.append(d_item)
            total_net += d_item["net"]
            total_gross += d_item["gross"]
            total_leave += d_item["leave"]
            if d_item["ot_declared"]:
                total_ot += d_item["overtime"]
            total_deficit += d_item["deficit"]
            if d_item["late"] > 0:
                total_late += d_item["late"]
                late_days += 1
            if d_item["work_mode"] == "remote":
                remote_days += 1

    return {
        "month_key": month_key,
        "month_name": m_name,
        "jy": jy,
        "jm": jm,
        "rows": rows,
        "net": total_net,
        "gross": total_gross,
        "leave": total_leave,
        "overtime": total_ot,
        "deficit": total_deficit,
        "late_total": total_late,
        "late_days": late_days,
        "work_days": work_days,
        "holiday_days": holiday_days,
        "holiday_worked": holiday_worked,
        "remote_days": remote_days,
        "dl_rows": dl_rows,
        "u_settings": u_settings,
    }

def get_month_report(conn: DBAdapter, month_key: str | None = None, user_id: int | None = None) -> dict:
    mk = month_key or current_month_key()
    m = compute_month(conn, mk, user_id=user_id)
    
    # Process daily leaves in-memory from batch
    dl_rows = m.get("dl_rows", [])
    dl_by_date = {}
    for r in dl_rows:
        for s in m.get("rows", []):
            sdate = s["date"]
            if r["start_date"] <= sdate <= r["end_date"]:
                dl_by_date[sdate] = r

    summary = Counter()
    workday_dl = 0
    for s, dl in dl_by_date.items():
        summary[dl["type"]] += 1
        is_hol, _ = record_service.holiday_name(conn, s)
        if not is_hol:
            workday_dl += 1

    m_leave_adj = m["leave"] + workday_dl * 8.0
    jy = m["jy"]
    quota = float(m["u_settings"].get("leave_quota_hours", "208") or 208.0)
    
    # Calculate yearly hourly leave in a single aggregated query
    if user_id is None:
        yr_events = conn.execute("SELECT ts_utc, event_type, note, shamsi_date FROM events WHERE substr(shamsi_date,1,4)=? AND user_id IS NULL ORDER BY ts_utc ASC", (f"{jy:04d}",)).fetchall()
    else:
        yr_events = conn.execute("SELECT ts_utc, event_type, note, shamsi_date FROM events WHERE substr(shamsi_date,1,4)=? AND user_id=? ORDER BY ts_utc ASC", (f"{jy:04d}", user_id)).fetchall()

    hourly_consumed = 0.0
    cur_ls = None
    for r in yr_events:
        et = r["event_type"]
        dt = datetime.datetime.fromisoformat(r["ts_utc"])
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        dt_tehran = dt.astimezone(settings.tehran_tz)
        if et == "leave_start":
            cur_ls = dt_tehran
        elif et == "leave_end" and cur_ls is not None:
            hourly_consumed += (record_service.company_clock(dt_tehran) - record_service.company_clock(cur_ls)).total_seconds() / 3600.0
            cur_ls = None

    annual_daily = leave_service.daily_leave_annual_hours_in_year(conn, user_id, jy) if user_id is not None else 0.0
    consumed_total = hourly_consumed + annual_daily

    leave_bal_info = {
        "quota": round(quota, 2),
        "consumed": round(consumed_total, 2),
        "remaining": round(max(0.0, quota - consumed_total), 2),
        "hourly": round(hourly_consumed, 2),
        "daily_annual": round(annual_daily, 2),
    }

    totals = {
        "net": round(m["net"], 2),
        "gross": round(m["gross"], 2),
        "leave": round(m_leave_adj, 2),
        "overtime": round(m["overtime"], 2),
        "deficit": round(m["deficit"], 2),
        "late_total": round(m["late_total"], 2),
        "late_days": m["late_days"],
        "work_days": m["work_days"],
        "holiday_days": m["holiday_days"],
        "holiday_worked": m["holiday_worked"],
        "remote_days": m["remote_days"],
    }

    report_text = f"گزارش ماه {m['month_name']} {m['jy']}\nکارکرد خالص: {totals['net']} ساعت\nکسری کار: {totals['deficit']} ساعت\nاضافه کار: {totals['overtime']} ساعت"

    return {
        "month_key": mk,
        "month_name": m["month_name"],
        "year": jy,
        "month": m["jm"],
        "rows": m["rows"],
        "totals": totals,
        "leave_balance": leave_bal_info,
        "daily_leaves_summary": dict(summary),
        "text": report_text,
    }

def get_week_report(conn: DBAdapter, user_id: int | None = None) -> dict:
    base = record_service.now_tehran()
    jy, jm, jd = jalali.gregorian_to_jalali(base.year, base.month, base.day)
    gy, gm, gd = jalali.jalali_to_gregorian(jy, jm, jd)
    today_date = datetime.date(gy, gm, gd)
    start_date = today_date - datetime.timedelta(days=(today_date.weekday() - 5) % 7)
    
    days = []
    total_net = total_ot = total_def = total_leave = 0.0
    work_days = 0
    remote_days = 0

    for i in range(7):
        cur_d = start_date + datetime.timedelta(days=i)
        cy, cm, cd = jalali.gregorian_to_jalali(cur_d.year, cur_d.month, cur_d.day)
        sdate = jalali.jalali_date_str(cy, cm, cd)
        d = record_service.day_payload(conn, sdate, user_id=user_id)
        if d.get("has_events") or (d.get("net") and d.get("net") > 0) or d.get("in"):
            days.append(d)
            total_net += d.get("net", 0.0)
            total_ot += d.get("overtime", 0.0)
            total_def += d.get("deficit", 0.0)
            total_leave += d.get("leave", 0.0)
            if not d.get("is_holiday"):
                work_days += 1
            if d.get("work_mode") == "remote":
                remote_days += 1

    totals = {
        "net": round(total_net, 2),
        "overtime": round(total_ot, 2),
        "deficit": round(total_def, 2),
        "leave": round(total_leave, 2),
        "work_days": work_days,
        "remote_days": remote_days,
    }

    report_text = f"گزارش هفته جاری\nکارکرد خالص: {totals['net']} ساعت\nکسری کار: {totals['deficit']} ساعت\nاضافه کار: {totals['overtime']} ساعت"

    return {
        "days": days,
        "totals": totals,
        "text": report_text,
    }

def export_excel(conn: DBAdapter, month_key: str | None = None, user_id: int | None = None) -> str:
    m = get_month_report(conn, month_key, user_id=user_id)
    wb = Workbook()
    ws = wb.active
    ws.title = m["month_name"]

    # Excel Styling
    f_title = Font(name="Tahoma", size=14, bold=True, color="FFFFFF")
    f_head = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
    f_data = Font(name="Tahoma", size=10)
    f_tot = Font(name="Tahoma", size=10, bold=True)
    
    fill_dark = PatternFill("solid", fgColor="0F172A")
    fill_amber = PatternFill("solid", fgColor="F59E0B")
    fill_row_alt = PatternFill("solid", fgColor="F8FAFC")
    fill_hol = PatternFill("solid", fgColor="FEF2F2")
    
    bd_thin = Side(style="thin", color="CBD5E1")
    border_cell = Border(left=bd_thin, right=bd_thin, top=bd_thin, bottom=bd_thin)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Header
    ws.merge_cells("A1:J1")
    tcell = ws["A1"]
    tcell.value = f"گزارش عملکرد ماهانه شیفت — {m['month_name']} {m['year']}"
    tcell.font = f_title
    tcell.fill = fill_dark
    tcell.alignment = align_center
    ws.row_dimensions[1].height = 40

    headers = [
        "ردیف", "تاریخ شمسی", "روز هفته", "وضعیت", "ورود", "خروج", "کارکرد (ساعت)", "کسری کار", "اضافه‌کار", "مرخصی"
    ]
    
    ws.append(headers)
    ws.row_dimensions[2].height = 26
    for col_idx in range(1, 11):
        c = ws.cell(row=2, column=col_idx)
        c.font = f_head
        c.fill = fill_amber
        c.alignment = align_center
        c.border = border_cell

    # Rows
    for idx, r in enumerate(m["rows"], start=1):
        row_data = [
            idx,
            r["date"],
            r["weekday"],
            r["work_mode_label"] if not r["is_holiday"] else (r["holiday_name"] or "تعطیل"),
            r["in"] or "—",
            r["out"] or "—",
            r["net"],
            r["deficit"],
            r["overtime"],
            r["leave"],
        ]
        ws.append(row_data)
        curr_row = idx + 2
        ws.row_dimensions[curr_row].height = 20
        for col_idx in range(1, 11):
            c = ws.cell(row=curr_row, column=col_idx)
            c.font = f_data
            c.alignment = align_center
            c.border = border_cell
            if r["is_holiday"]:
                c.fill = fill_hol
            elif idx % 2 == 0:
                c.fill = fill_row_alt

    # Totals Row
    tot_row = len(m["rows"]) + 3
    ws.merge_cells(f"A{tot_row}:F{tot_row}")
    tc = ws.cell(row=tot_row, column=1)
    tc.value = "مجموع عملکرد ماهانه"
    tc.font = f_tot
    tc.alignment = align_right
    
    ws.cell(row=tot_row, column=7, value=m["totals"]["net"]).font = f_tot
    ws.cell(row=tot_row, column=8, value=m["totals"]["deficit"]).font = f_tot
    ws.cell(row=tot_row, column=9, value=m["totals"]["overtime"]).font = f_tot
    ws.cell(row=tot_row, column=10, value=m["totals"]["leave"]).font = f_tot

    for col_idx in range(1, 11):
        ws.cell(row=tot_row, column=col_idx).border = border_cell

    ws.views.sheetView[0].rightToLeft = True

    # Adjust widths
    col_widths = {1: 8, 2: 14, 3: 12, 4: 16, 5: 10, 6: 10, 7: 15, 8: 12, 9: 12, 10: 12}
    for c_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    os.makedirs(settings.EXPORTS_DIR, exist_ok=True)
    fpath = os.path.join(settings.EXPORTS_DIR, f"report_{m['month_key']}_{user_id or 'anon'}.xlsx")
    wb.save(fpath)
    return fpath
